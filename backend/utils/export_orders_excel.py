#!/usr/bin/env python3
"""
订单Excel导出脚本
使用openpyxl生成Excel文件，包含所有订单必要信息
Requirements: 8.4, 8.5
"""

import sys
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime


def export_orders_excel(orders, output_path=None):
    """
    将实体产品订单导出为Excel文件
    
    Args:
        orders: 订单数据列表 [{
            order_id, user_name, phone, address, 
            product_type, image_url, create_time
        }]
        output_path: 输出文件路径（可选）
        
    Returns:
        dict: {success: bool, output_path: str, order_count: int, message: str}
    """
    try:
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "实体产品订单"
        
        # 定义表头
        headers = [
            "订单编号",
            "用户姓名",
            "联系电话",
            "收货地址",
            "产品类型",
            "艺术照URL",
            "下单时间"
        ]
        
        # 设置表头样式
        header_font = Font(name='微软雅黑', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        # 边框样式
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 写入表头
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # 设置列宽
        column_widths = {
            'A': 20,  # 订单编号
            'B': 15,  # 用户姓名
            'C': 15,  # 联系电话
            'D': 40,  # 收货地址
            'E': 12,  # 产品类型
            'F': 50,  # 艺术照URL
            'G': 20   # 下单时间
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # 写入订单数据
        for row_idx, order in enumerate(orders, start=2):
            # 产品类型映射
            product_type_map = {
                'crystal': '晶瓷画',
                'scroll': '卷轴'
            }
            
            # 格式化时间
            create_time = order.get('create_time', '')
            if isinstance(create_time, str):
                try:
                    dt = datetime.fromisoformat(create_time.replace('Z', '+00:00'))
                    create_time = dt.strftime('%Y-%m-%d %H:%M:%S')
                except:
                    pass
            
            row_data = [
                order.get('order_id', ''),
                order.get('user_name', ''),
                order.get('phone', ''),
                order.get('address', ''),
                product_type_map.get(order.get('product_type', ''), order.get('product_type', '')),
                order.get('image_url', ''),
                create_time
            ]
            
            # 写入数据
            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                cell.border = thin_border
                
                # URL列使用蓝色字体
                if col_idx == 6:
                    cell.font = Font(name='微软雅黑', size=10, color='0563C1', underline='single')
        
        # 冻结首行
        ws.freeze_panes = 'A2'
        
        # 保存文件
        if output_path is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_path = f"product_orders_{timestamp}.xlsx"
        
        wb.save(output_path)
        
        return {
            'success': True,
            'output_path': output_path,
            'order_count': len(orders),
            'message': f'成功导出 {len(orders)} 条订单'
        }
    
    except Exception as e:
        return {
            'success': False,
            'message': f'Excel导出失败: {str(e)}'
        }


def main():
    """
    命令行入口
    接收JSON格式的参数: {
        "orders": [{
            "order_id": "...",
            "user_name": "...",
            "phone": "...",
            "address": "...",
            "product_type": "crystal|scroll",
            "image_url": "...",
            "create_time": "..."
        }],
        "output_path": "..."
    }
    """
    try:
        # 从命令行参数读取JSON
        if len(sys.argv) > 1:
            params = json.loads(sys.argv[1])
        else:
            # 从stdin读取
            params = json.load(sys.stdin)
        
        orders = params.get('orders', [])
        output_path = params.get('output_path')
        
        if not orders:
            result = {
                'success': False,
                'message': '缺少必需参数: orders'
            }
        else:
            result = export_orders_excel(orders, output_path)
        
        # 输出JSON结果
        print(json.dumps(result, ensure_ascii=False))
        
    except Exception as e:
        result = {
            'success': False,
            'message': f'脚本执行失败: {str(e)}'
        }
        print(json.dumps(result, ensure_ascii=False))
        sys.exit(1)


if __name__ == '__main__':
    main()
